import openpyxl

class XLread:
    @staticmethod
    def read_data(path,sheetn,row,column):
        wb=openpyxl.load_workbook(path)
        sheet = wb[sheetn]
        return sheet.cell(row=row,column=column).value
    @staticmethod
    def noofrows(path,sheetn):
        wb=openpyxl.load_workbook(path)
        sheet = wb[sheetn]
        return sheet.max_row
    @staticmethod
    def noofcolumns(path,sheetn):
        wb=openpyxl.load_workbook(path)
        sheet = wb[sheetn]
        return sheet.max_column